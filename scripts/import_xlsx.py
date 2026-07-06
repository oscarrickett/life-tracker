"""
Import all 'Template of my Life YYYY.xlsx' workbooks into a single seed.json.

Quirks handled:
- All workbooks use sheet name '2022' regardless of year.
- 'Template of my Life 2026.xlsx' has 2026 data but its A column still holds
  2025 dates (the user copied last year's file). Day-of-week is correct, so
  for the 2026 file we ignore col A and reconstruct the date from row index.
- The category key block (right of the data) drifts across years
  (e.g. category 5 changed meaning in 2025). We capture per-year keys.
  RENAMES below relabels historical names we don't want in the output.

Output:
    data/seed.json
        {
          "categories": {"1":"Sleep", ...},          # canonical, latest
          "categoriesByYear": {"2023": {...}, ...},  # historical
          "days": [
             {"date":"2023-01-01","day":"Sunday","hours":[1,1,...,18],"notes":"..."},
             ...
          ]
        }

Usage:
    python scripts/import_xlsx.py
"""
from __future__ import annotations

import datetime as dt
import json
import os
import sys
from pathlib import Path

import openpyxl

DESKTOP = Path(r"E:\Desktop")
OUT = Path(__file__).resolve().parent.parent / "data" / "seed.json"

WORKBOOKS = {
    2023: "Template of my Life 2023.xlsx",
    2024: "Template of my Life 2024.xlsx",
    2025: "Template of my Life 2025.xlsx",
    2026: "Template of my Life 2026.xlsx",
}

# Columns: A=date, B=day, C..Z = 24 hours (00..23), AA = notes (col 27)
HOUR_COL_START = 3   # 'C'
HOUR_COL_END = 26    # 'Z' inclusive
NOTES_COL = 27       # 'AA'

# Category key block is at columns AC (29) / AD (30) / AE (31) / AF (32),
# laid out as two pairs per row (id, name) (id, name) starting around row 3.
KEY_LEFT_ID, KEY_LEFT_NAME = 29, 30
KEY_RIGHT_ID, KEY_RIGHT_NAME = 31, 32

# Relabel historical category names on import (seed.json is public).
RENAMES = {"Clare": "Relationship"}


def read_category_key(ws) -> dict[int, str]:
    out: dict[int, str] = {}
    for r in range(2, 25):
        for id_col, name_col in ((KEY_LEFT_ID, KEY_LEFT_NAME), (KEY_RIGHT_ID, KEY_RIGHT_NAME)):
            cid = ws.cell(row=r, column=id_col).value
            name = ws.cell(row=r, column=name_col).value
            if isinstance(cid, (int, float)) and name:
                clean = str(name).strip()
                out[int(cid)] = RENAMES.get(clean, clean)
    return out


def read_category_colors(ws) -> dict[int, str]:
    """Pull the conditional-formatting color per category id ('cellIs == N')."""
    out: dict[int, str] = {}
    for _rng, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            if rule.type != "cellIs" or rule.operator != "equal":
                continue
            try:
                cid = int(float(rule.formula[0]))
            except (ValueError, IndexError, TypeError):
                continue
            df = getattr(rule, "dxf", None)
            fill = getattr(df, "fill", None) if df else None
            if not fill:
                continue
            bg = getattr(fill, "bgColor", None)
            rgb = getattr(bg, "rgb", None) if bg else None
            if isinstance(rgb, str) and len(rgb) == 8:  # 'AARRGGBB'
                out[cid] = "#" + rgb[2:].upper()
    return out


def import_year(path: Path, year: int) -> tuple[list[dict], dict[int, str], dict[int, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2022"]
    key = read_category_key(ws)
    colors = read_category_colors(ws)

    days: list[dict] = []
    jan1 = dt.date(year, 1, 1)
    for r in range(2, ws.max_row + 1):
        # Reconstruct date from row index for 2026 (date col is wrong),
        # otherwise trust column A.
        if year == 2026:
            date = jan1 + dt.timedelta(days=r - 2)
            if date.year != year:
                break
        else:
            v = ws.cell(row=r, column=1).value
            if not hasattr(v, "year"):
                continue
            date = dt.date(v.year, v.month, v.day)
            if date.year != year:
                continue

        day_name = ws.cell(row=r, column=2).value
        hours = []
        for c in range(HOUR_COL_START, HOUR_COL_END + 1):
            cv = ws.cell(row=r, column=c).value
            hours.append(int(cv) if isinstance(cv, (int, float)) else None)

        notes = ws.cell(row=r, column=NOTES_COL).value
        notes = str(notes).strip() if notes else ""

        # Skip rows that are entirely empty for 2026 (future days).
        if all(h is None for h in hours) and not notes:
            if year == 2026:
                continue

        days.append({
            "date": date.isoformat(),
            "day": str(day_name) if day_name else "",
            "hours": hours,
            "notes": notes,
        })
    return days, key, colors


def main() -> int:
    all_days: list[dict] = []
    keys_by_year: dict[str, dict[str, str]] = {}
    colors_by_year: dict[str, dict[str, str]] = {}
    for year, fn in WORKBOOKS.items():
        path = DESKTOP / fn
        if not path.exists():
            print(f"missing: {path}", file=sys.stderr)
            continue
        days, key, colors = import_year(path, year)
        all_days.extend(days)
        keys_by_year[str(year)] = {str(k): v for k, v in sorted(key.items())}
        colors_by_year[str(year)] = {str(k): v for k, v in sorted(colors.items())}
        filled = sum(1 for d in days if any(h is not None for h in d["hours"]))
        print(f"{year}: {len(days)} day-rows, {filled} with hours, "
              f"{len(key)} categories, {len(colors)} colored")

    canonical: dict[str, str] = {}
    for year_str in sorted(keys_by_year.keys(), reverse=True):
        for cid, name in keys_by_year[year_str].items():
            canonical.setdefault(cid, name)

    # Canonical color = 2026's color if defined, else fall back to any earlier year
    canonical_colors: dict[str, str] = {}
    for year_str in sorted(colors_by_year.keys(), reverse=True):
        for cid, col in colors_by_year[year_str].items():
            canonical_colors.setdefault(cid, col)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "categories": canonical,
        "categoriesByYear": keys_by_year,
        "colors": canonical_colors,
        "colorsByYear": colors_by_year,
        "days": all_days,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT} ({OUT.stat().st_size:,} bytes, {len(all_days)} days)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
